import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
import base64
import plotly.express as px  # Import Plotly Express
import plotly.graph_objects as go  # Import graph_objects
import io
import os
import matplotlib.pyplot as plt
import numpy as np

def go_to_step(step):
        """Callback function to update the current step."""
        st.session_state["current_step"] = step


# --- Sample Project Data (Replace with your actual data) ---
project_data = {
    "Portfolio": [
        "Portfolio A",
        "Portfolio A",
        "Portfolio A",
        "Portfolio B",
        "Portfolio B",
    ],
    "Sub-Portfolio": [
        "Sub-Portfolio 1",
        "Sub-Portfolio 2",
        "Sub-Portfolio 3",
        "Sub-Portfolio 4",
        "Sub-Portfolio 5",
    ],
    "Project Name": [
        "Project 1",
        "Project 2",
        "Project 3",
        "Project 4",
        "Project 5",
    ],
}
project_df = pd.DataFrame(project_data)

# --- Functions ---

def chatbot_message(message):
    """Displays a chatbot message."""
    st.markdown(
        f"""
        <div class="chatbot-message">
            <div class="chatbot-avatar">
                <span>🤖</span>
            </div>
            <div class="chatbot-text">
                {message}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def user_message(message):
    """Displays a user message."""
    st.markdown(
        f"""
        <div class="user-message">
            <div class="user-text">
                {message}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def save_project_plan(project_name, description, stakeholders, target_end_date, milestones, file_path):
    """Saves the project plan to an Excel file using the browser's 'Save As' dialog."""

    try:
        wb = openpyxl.Workbook()
        ws = wb.active

        ws["A1"] = "Project Name:"
        ws["B1"].font = Font(size=14, bold=True)
        ws["A2"] = "Description:"
        ws["A3"] = "Stakeholders:"
        ws["A4"] = "Target End Date:"
        ws["A6"] = "Milestone Name"
        ws["B6"] = "Start Date"
        ws["C6"] = "End Date"
        ws["D6"] = "Milestone Owner"
        ws["E6"] = "Progress"
        ws["F6"] = "Status"

        ws["B1"] = project_name
        ws["B2"] = description
        ws["B3"] = stakeholders
        ws["B4"] = target_end_date.strftime("%Y-%m-%d") if target_end_date else "N/A"

        for i, milestone in enumerate(milestones, start=7):
            ws[f"A{i}"] = milestone["Name"]
            ws[f"B{i}"] = milestone["Start Date"].strftime("%Y-%m-%d") if milestone["Start Date"] else "N/A"
            ws[f"C{i}"] = milestone["End Date"].strftime("%Y-%m-%d") if milestone["End Date"] else "N/A"
            ws[f"D{i}"] = milestone["Milestone Owner"]
            ws[f"E{i}"] = milestone["Progress"]
            ws[f"F{i}"] = milestone.get("Status", "")

        # Save to the current working directory
        full_file_path = os.path.join(os.getcwd(), file_path)
        wb.save(full_file_path)
        # st.success(f"Project plan saved as '{file_path}' in '{os.getcwd()}'")

    except Exception as e:
        st.error(f"Error saving the project plan: {e}")


def load_project_plan(file_path):
    """Loads project plan data from an Excel file, handling 'N/A' for Target End Date."""
    wb = openpyxl.load_workbook(file_path, data_only=True) # data_only=True to read cell values
    ws = wb.active

    project_name = ws["B1"].value
    description = ws["B2"].value
    stakeholders = ws["B3"].value

    target_end_date_str = ws["B4"].value
    if target_end_date_str == "N/A" or target_end_date_str is None:
        target_end_date = None
    else:
        try:
            target_end_date = datetime.strptime(str(target_end_date_str), "%Y-%m-%d").date()  # Convert to string before parsing
        except ValueError:
            st.error(
                f"Invalid date format in the Excel file for 'Target End Date'. Please use YYYY-MM-DD format."
            )
            return None, None, None, None, None
        
    # Handle target_end_date separately as it might be a datetime object already
    target_end_date_str = ws["B4"].value
    if target_end_date_str == "N/A" or target_end_date_str is None:
        target_end_date = None
    elif isinstance(target_end_date_str, datetime):  # Check if it's already a datetime object
        target_end_date = target_end_date_str.date()
    else:
        target_end_date = datetime.strptime(target_end_date_str, "%Y-%m-%d").date()

    milestones = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] is None:
            break
        name, start_date, end_date, owner, progress, status_progress = row

        # Convert to datetime.date, handling potential 'N/A' values
        if isinstance(start_date, datetime):
            start_date = start_date.date() 
        elif isinstance(start_date, str) and start_date != 'N/A':
            try:
                start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            except ValueError:
                start_date = None # Or handle the invalid date format appropriately
        else:
            start_date = None 

        if isinstance(end_date, datetime):
            end_date = end_date.date() 
        elif isinstance(end_date, str) and end_date != 'N/A':
            try:
                end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                end_date = None # Or handle the invalid date format appropriately
        else:
            end_date = None

        milestones.append(
            {
                "Name": name,
                "Start Date": start_date,  
                "End Date": end_date,    
                "Milestone Owner": owner,
                "Progress": progress,
                "Status": status_progress
            }
        )

    return project_name, description, stakeholders, target_end_date, milestones


def display_project_details(
    project_name, description, stakeholders, target_end_date, milestones
):
    """Displays the loaded project details."""
    st.subheader("Loaded Project Details:")
    st.write(f"Project Name: {project_name}")
    st.write(f"Description: {description}")
    st.write(f"Stakeholders: {stakeholders}")
    st.write(
        f"Target End Date: {target_end_date.strftime('%Y-%m-%d') if target_end_date else 'N/A'}"
    )

    st.subheader("Milestones:")
    milestones_display = [
        {
            "Name": milestone["Name"],
            "Start Date": milestone["Start Date"].strftime("%Y-%m-%d") if milestone["Start Date"] else "N/A",
            "End Date": milestone["End Date"].strftime("%Y-%m-%d") if milestone["End Date"] else "N/A",
            "Milestone Owner": milestone["Milestone Owner"],
            "Progress": milestone["Progress"],
            "Status": milestone.get("Status", "")
        }
        for milestone in milestones
    ]
    df = pd.DataFrame(milestones_display)
    st.dataframe(df, hide_index=True)

import streamlit as st
import plotly.graph_objects as go  # Make sure you have Plotly installed!
import pandas as pd


import streamlit as st
import plotly.graph_objects as go
import pandas as pd

def display_plotly_progress(milestone_data):
        st.subheader("Milestone Progress Visualization:")

        df = pd.DataFrame(milestone_data)
        milestone_names = df["Name"].tolist()

        fig = go.Figure()

        # --- TOTAL PROGRESS BAR ---
        fig.add_trace(
            go.Bar(
                x=milestone_names,
                y=[100] * len(milestone_names),
                marker_color="lightgray",
                name="Total",  # Name for the legend
                marker_line_color="white",
                marker_line_width=2,
            )
        )

        annotations = []  
        for i, row in df.iterrows():
            actual_progress = float(row["Actual Progress"])
            expected_progress = float(row["Expected Progress"])

            # --- EXPECTED PROGRESS (Orange) ---
            if expected_progress > actual_progress:
                fig.add_trace(
                    go.Bar(
                        x=[row["Name"]],  
                        y=[expected_progress],
                        marker_color="darkorange",
                        name="Expected Progress", 
                        width=0.6, 
                        showlegend=False # Hide from legend 
                    )
                )

                # Annotation for Expected Progress (above)
                # Format the Expected Progress text:
                if expected_progress == 100.00:
                    expected_text = "100%"
                else:
                    expected_text = f"{expected_progress:.2f}%"
    
                annotations.append(
                    dict(
                        x=row["Name"],
                        y=expected_progress + 5,
                        text=expected_text,  # Use the formatted text here
                        showarrow=False,
                        font=dict(color="black"),
                    )
                )
            # --- ACTUAL PROGRESS (Green) --- 
            fig.add_trace(
                go.Bar(
                    x=[row["Name"]], 
                    y=[actual_progress],
                    marker_color="green",
                    name="Actual Progress",  
                    width=0.6,
                    showlegend=False  # Hide from legend if expected_progress is shown
                )
            )

            # Annotation for Actual Progress (below or center)
            if expected_progress > actual_progress:
                y_position = actual_progress - 5  # Below orange bar
            else:
                y_position = actual_progress / 2  # Centered

            annotations.append(
                dict(
                    x=row["Name"],
                    y=y_position, 
                    text=f"{actual_progress:.0f}%" if actual_progress < 100 else "",
                    showarrow=False,
                    font=dict(color="white"),
                )
            )

        # Update the figure layout
        fig.update_layout(
            barmode="overlay",
            xaxis_tickangle=-45,
            title="Milestone Progress",
            xaxis_title="Milestone Name",
            yaxis_title="Progress (%)",
            plot_bgcolor="rgba(0,0,0,0)",  # Transparent background
            annotations=annotations,  # Add annotations to the layout
            bargap=0.1,
            yaxis=dict(
                range=[
                    0,
                    max(
                        [float(row["Expected Progress"]) for _, row in df.iterrows()],
                        default=100,
                    )
                    + 10,
                ]
            ),
        )

        # Render the chart!
        st.plotly_chart(fig)

def milestone_actions(milestones, file_path=None):
    """Handles adding milestones with confirmation."""
    st.subheader("Milestone Actions:")

    col1, col2 = st.columns(2) # Create two columns

    with col1:
    # --- Milestone Modification Controls (Column 1) ---
        action = st.radio(
            "Choose an action:",
            ("Add New Milestone", "Modify Existing Milestone", "Delete Milestone"),
            key="milestone_action_radio",
        )


        if action == "Add New Milestone":
            st.write("**Adding New Milestone:**")
            new_milestone = {
                "Name": st.text_input("Milestone Name"),
                "Start Date": st.date_input("Start Date"),  # Don't convert to string here
                "End Date": st.date_input("End Date"),    # Don't convert to string here
                "Milestone Owner": st.text_input("Milestone Owner"),
                "Progress": st.number_input("Progress (0-100)", min_value=0, max_value=100, step=1),
                "Status": st.text_input("Status"),  
            }
            if st.button("Add Milestone"):
                milestones.append(new_milestone)
                st.success("Milestone added!")
                

        elif action == "Modify Existing Milestone":
            milestone_to_modify = st.selectbox(
                "Select Milestone to Modify",
                [m["Name"] for m in milestones],
                key="modify_milestone_selectbox",
            )
            st.session_state["selected_milestone"] = milestone_to_modify
            for i, milestone in enumerate(milestones):
                if milestone["Name"] == milestone_to_modify:
                    st.write(f"**Modifying Milestone:** {milestone_to_modify}")
                    milestones[i]["Name"] = st.text_input(
                        "Milestone Name", value=milestone["Name"]
                    )

                    # Convert string dates to datetime objects:
                    start_date_value = (
                        datetime.strptime(milestone["Start Date"], "%Y-%m-%d").date()
                        if isinstance(milestone["Start Date"], str)
                        else milestone["Start Date"]
                    )
                    end_date_value = (
                        datetime.strptime(milestone["End Date"], "%Y-%m-%d").date()
                        if isinstance(milestone["End Date"], str)
                        else milestone["End Date"]
                    )

                    milestones[i]["Start Date"] = st.date_input(
                        "Start Date", value=start_date_value
                    )  # No need to convert to string

                    milestones[i]["End Date"] = st.date_input(
                        "End Date", value=end_date_value
                    )  # No need to convert to string

                    milestones[i]["Milestone Owner"] = st.text_input(
                        "Milestone Owner", value=milestone["Milestone Owner"]
                    )
                    milestones[i]["Progress"] = st.number_input(
                        "Progress (0-100)",
                        min_value=0,
                        max_value=100,
                        value=int(milestone["Progress"])
                        if isinstance(milestone["Progress"], int)
                        else 0,
                        step=1,
                    )

                    # Modify Status  in the UI
                    milestones[i]["Status"] = st.text_input(
                        "Status",
                        value=milestone.get("Status", ""),
                    )

                    break
            if st.button("Save Milestone Changes"):
                st.success("Milestone modified!")
                

        elif action == "Delete Milestone":
            milestone_to_delete = st.selectbox(
                "Select Milestone to Delete",
                [m["Name"] for m in milestones],
                key="delete_milestone_selectbox",
            )
            st.session_state["selected_milestone"] = milestone_to_delete
            if st.button("Confirm Delete"):
                for i, milestone in enumerate(milestones):
                    if milestone["Name"] == milestone_to_delete:
                        del milestones[i]
                        st.success(f"Milestone '{milestone_to_delete}' deleted!")
                        break
                    

    with col2:
        # --- Live Project Plan Preview (Column 2) ---
        st.subheader("Project Plan Preview:")

        # Display the loaded project details dynamically:
        display_project_details(
            st.session_state["project_name"],
            st.session_state["description"],
            st.session_state["stakeholders"],
            st.session_state["target_end_date"],
            st.session_state["milestones"], # Display the updated milestones 
        )

    # Add the "Next" button after the modification controls
    if st.button("Next",on_click=go_to_step, args=(6,), key="next_button_5"):
        # Save the modified project plan
        try:
            if "file_path" in st.session_state and st.session_state["file_path"]:
                save_project_plan(
                    st.session_state["project_name"],
                    st.session_state["description"],
                    st.session_state["stakeholders"],
                    st.session_state["target_end_date"],
                    st.session_state["milestones"],
                    st.session_state["file_path"],  # Use the original file name
                )

                st.success(f"Project plan saved to '{st.session_state['file_path']}'!")

                
                st.session_state["current_step"] = 12  # Proceed to Step 12
                


                # # Prompt to analyze or provide feedback after saving
                # chatbot_message(
                #     "Your modified project plan is ready! What would you like to do next?"
                # )
                # col1, col2 = st.columns(2)
                # with col1:
                #     if st.button("Analyse it"):
                #         st.session_state["current_step"] = 7.1  # Analyze modified plan
                # with col2:
                #     if st.button("Provide Feedback"):
                #         st.session_state["current_step"] = 13  # Feedback 

            else:
                # st.error("No file path found. Please upload a file first.")
                st.session_state["current_step"] = 11  # Go back to file upload
        except Exception as e:
            st.error(f"Error saving: {e}")

        

    return milestones
        

# --- Milestone Analysis Functions ---

def calculate_program_duration(milestones):
    """Calculates the extreme start and end dates and the program duration."""
    start_dates = [m["Start Date"] for m in milestones]  # No need for strptime()
    end_dates = [m["End Date"] for m in milestones]  # No need for strptime()

    # Convert to datetime.date objects if needed:
    start_dates = [d.date() if isinstance(d, datetime) else d for d in start_dates]
    end_dates = [d.date() if isinstance(d, datetime) else d for d in end_dates]

    extreme_start_date = min(start_dates)
    extreme_end_date = max(end_dates)
    program_duration = (extreme_end_date - extreme_start_date).days

    return extreme_start_date, extreme_end_date, program_duration


def analyze_milestone_progress(milestone, extreme_start_date, extreme_end_date, target_end_date):
    """Analyzes the progress of a single milestone."""
    milestone_start = milestone["Start Date"] 
    milestone_end = milestone["End Date"] 
    progress = int(milestone["Progress"]) 

    milestone_start = milestone_start.date() if isinstance(milestone_start, datetime) else milestone_start
    milestone_end = milestone_end.date() if isinstance(milestone_end, datetime) else milestone_end

    today = datetime.now().date()

    # 1. Calculate Expected Progress Based on Time:
    total_milestone_days = (milestone_end - milestone_start).days
    days_elapsed = (today - milestone_start).days

    expected_progress = (days_elapsed / total_milestone_days) * 100 if total_milestone_days > 0 else 100
    expected_progress = min(expected_progress, 100)  # Cap at 100%

    # 2. Determine Milestone Status: 
    status = "On Track"  # Default status
    warning = "" 

    if milestone_end < today:  # Milestone end date has passed
        if progress == 100:
            status = "Completed"
        else:
            status = "Behind Schedule"
            warning = f"Milestone '{milestone['Name']}' is past its end date and not 100% complete."
    elif milestone_start <= today <= milestone_end:  # Milestone is in progress
        if progress < expected_progress - 10:
            status = "Behind Schedule"
            warning = f"Milestone '{milestone['Name']}' is behind schedule! It should be at about {expected_progress:.2f}% progress."
    
    # Check against target end date (if provided)
    if target_end_date and milestone_end > target_end_date:
        status = "Critical"
        warning = f"Milestone '{milestone['Name']}' end date is after the project target end date! This needs immediate attention."

    return status, warning, expected_progress


# --- Dependency Functions ---

def add_dependency_milestone(milestones, dependent_project_name, start_date, end_date, progress=0):  # Add progress parameter
    """Adds a dependency milestone to the project plan."""
    dependency_milestone_name = f"d - Dependency on {dependent_project_name}"
    new_milestone = {
        "Name": dependency_milestone_name,
        "Start Date": start_date, # No need for strftime here (dates are already datetime.date objects)
        "End Date": end_date,   # No need for strftime here
        "Milestone Owner": "N/A",
        "Progress": progress,     # Use the provided progress value
        "Status": "",
    }
    milestones.append(new_milestone)
    return milestones

def handle_dependency_input(milestones, project_name, file_path):
    """Handles user input for dependency details."""
    chatbot_message("Please select the portfolio and sub-portfolio for the dependent project.")
    portfolio = st.selectbox("Portfolio", project_df["Portfolio"].unique())
    sub_portfolio = st.selectbox("Sub-Portfolio", project_df[project_df["Portfolio"] == portfolio]["Sub-Portfolio"].unique())

    # Filter projects based on selected portfolio and sub-portfolio
    filtered_projects = project_df[
        (project_df["Portfolio"] == portfolio) & (project_df["Sub-Portfolio"] == sub_portfolio)
    ]["Project Name"].unique()

    dependent_project_name = st.selectbox(
        "Dependent Project Plan Name:", filtered_projects
    )
    start_date = st.date_input("Dependency Start Date", value=None)
    end_date = st.date_input("Dependency End Date", value=None)
    progress = st.number_input("Dependency Progress (0-100)", min_value=0, max_value=100, value=0) # Add Progress input

    # Add dependency milestone to the current project
    if st.button(f"Add Dependency") and dependent_project_name and start_date and end_date:
        milestones = add_dependency_milestone(
            milestones,
            dependent_project_name,
            start_date,
            end_date,
            progress  # Pass progress to the function
        )

        # --- Do NOT save the plan here. Wait for the "Next" button ---
        st.success(f"Dependency milestone added to '{project_name}'!")

        # --- Add "Next" button for saving ---
        if st.button("Next", on_click=go_to_step, args=(6,), key="next_dependency_button"):
            try:

        # Save the current project plan
                save_project_plan(
                    st.session_state["project_name"],
                    st.session_state["description"],
                    st.session_state["stakeholders"],
                    st.session_state["target_end_date"],
                    milestones,
                    file_path,
                )
                st.success(f"Project plan saved to '{file_path}'!")

                if not file_path:
                    st.session_state["current_step"] = 7  # Go back to file upload

                else:  # If modified project
                    st.session_state["current_step"] = 12  # Go to analyze/feedback options
            except Exception as e:
                st.error(f"Error saving: {e}")

    else:
        st.warning("Please fill in all the dependent project plan details before adding the milestone.")

        


# --- Streamlit UI ---
st.set_page_config(page_title="Project Plan Chatbot", page_icon="🤖")

st.title("Chatbot Project Plan Creator & Analyzer")

# CSS Styling
st.markdown(
    """
    <style>
        .chatbot-message {
            display: flex;
            align-items: flex-start;
            margin-bottom: 15px;
        }

        .chatbot-avatar {
            background-color: #f0f0f0;
            padding: 10px;
            border-radius: 50%;
            margin-right: 10px;
        }

        .chatbot-avatar span {
            font-size: 20px;
        }

        .chatbot-text {
            background-color: #e5e5e5;
            padding: 10px;
            border-radius: 5px;
        }

        .user-message {
            display: flex;
            align-items: flex-end;
            justify-content: flex-end;
            margin-bottom: 15px;
        }

        .user-text {
            background-color: #c2f0c2;
            padding: 10px;
            border-radius: 5px;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# Session State Initialization
if "current_step" not in st.session_state:
    st.session_state["current_step"] = 0
if "num_milestones" not in st.session_state:
    st.session_state["num_milestones"] = 0
if "milestone_index" not in st.session_state:
    st.session_state["milestone_index"] = 0
if "project_name" not in st.session_state:
    st.session_state["project_name"] = ""
if "description" not in st.session_state:
    st.session_state["description"] = ""
if "stakeholders" not in st.session_state:
    st.session_state["stakeholders"] = ""
if "target_end_date" not in st.session_state:
    st.session_state["target_end_date"] = None
if "milestones" not in st.session_state:
    st.session_state["milestones"] = []
if "rating" not in st.session_state:
    st.session_state["rating"] = 1
if "suggestions" not in st.session_state:
    st.session_state["suggestions"] = ""
if "file_path" not in st.session_state:
    st.session_state["file_path"] = ""
if "milestones_entered" not in st.session_state:
    st.session_state["milestones_entered"] = False
if "dependency_input" not in st.session_state:
    st.session_state["dependency_input"] = False

# for key in [
#     "current_step",
#     "num_milestones",
#     "milestone_index",
#     "project_name",
#     "description",
#     "stakeholders",
#     "target_end_date",
#     "milestones",
#     "rating",
#     "suggestions",
#     "file_path",
#     "milestones_entered",
#     "dependency_input",
# ]:
#     if key not in st.session_state:
#         st.session_state[key] = "" if key != "milestones" else []


# ... (rest of your code)

# --- File Upload Section ---
def file_upload_section():
    """Handles file upload and data loading, branching to modify or analyze."""
    chatbot_message("Please upload your project plan (.xlsx file).")
    uploaded_file = st.file_uploader("Choose a file", type=["xlsx"])

    if uploaded_file is not None:
        try:
            (
                st.session_state["project_name"],
                st.session_state["description"],
                st.session_state["stakeholders"],
                st.session_state["target_end_date"],
                st.session_state["milestones"],
            ) = load_project_plan(uploaded_file.name)

            if all(
                val is not None
                for val in [
                    st.session_state["project_name"],
                    st.session_state["description"],
                    st.session_state["stakeholders"],
                    st.session_state["milestones"],
                ]
            ):
                chatbot_message("Project plan loaded successfully!")
                st.session_state["file_path"] = uploaded_file.name

                # Show the correct button based on the initial choice
                if st.session_state["current_step"] == 2:  # Analyze
                    if st.button("Let's analyze it!"):
                        st.session_state["current_step"] = 7.1  # Analyze loaded plan
                elif st.session_state["current_step"] == 11:  # Modify
                    if st.button("Let's modify it!"):
                        st.session_state["current_step"] = 5  # Modify loaded plan

            else:
                st.error("Error loading. Check file format.")
                st.session_state["current_step"] = 2  

        except Exception as e:
            st.error(f"Error loading: {e}")
            st.session_state["current_step"] = 2 

# ... (rest of your code)

# --- Chatbot Flow ---
def chatbot_flow():
    current_step = st.session_state.get("current_step", 0)

    # def go_to_step(step):
    #     """Callback function to update the current step."""
    #     st.session_state["current_step"] = step

    if current_step == 0:
        chatbot_message(
            "Welcome! I'm your friendly project planning chatbot. How can I help you today?"
        )
        chatbot_message(
            "Do you want to: \n Create a new project plan or\n Analyze an existing project plan or\n Modify an existing project plan"
        )

        col1, col2, col3 = st.columns(3)  # Three columns for buttons
        with col1:
            st.button("Create New Plan", on_click=go_to_step, args=(3,)) # Use callback
        with col2:
            st.button("Analyze Existing Plan", on_click=go_to_step, args=(2,))
        with col3:
            st.button("Modify Existing Plan", on_click=go_to_step, args=(11,))

    elif current_step == 2:  # Analyze Existing Project Plan
        chatbot_message("Great! Let's analyze your existing project plan.")
        file_upload_section()  # This function should handle the file upload and loading

        # Directly proceed to project analysis if a file is uploaded
        if st.session_state["file_path"]:  
            st.session_state["current_step"] = 7.1  # Analyze loaded plan

    elif current_step == 3:  # New Project Creation
        chatbot_message("Great! Let's get started.")

        st.session_state["project_name"] = st.text_input(
            "What is the name of your project?", value=st.session_state["project_name"]
        )
        st.session_state["description"] = st.text_area(
            "Can you describe your project briefly?",
            value=st.session_state["description"],
        )
        st.session_state["stakeholders"] = st.text_input(
            "Who are the key stakeholders involved in this project?",
            value=st.session_state["stakeholders"],
        )
        st.session_state["target_end_date"] = st.date_input(
            "What is the target end date for your project?",
            value=st.session_state["target_end_date"],
        )

        st.button("Next", on_click=go_to_step, args=(3.1,))

    elif current_step == 3.1:  # Number of Milestones Input
        try:
            num_milestones = int(
                st.text_input("How many milestones would you like to set up?")
            )
            if num_milestones <= 0:
                st.warning("Please enter a positive integer.")
            else: 
                st.session_state["num_milestones"] = num_milestones # Moved this line BEFORE the button
                if st.button("Next", on_click=go_to_step, args=(4,), key="next_button_3_1"):
                    st.session_state["milestone_index"] = 0  # Reset the milestone index
                    # ... (rest of your code) 
        except ValueError:
            st.warning("Please enter a valid integer.") 
        # else:
        #     st.warning("Please enter the number of milestones.")
        
    # elif current_step == 4:  # Milestone Input
    #     print("Current Step:", current_step)
    #     print("Milestone Index:", st.session_state["milestone_index"]) 
    #     print("Number of Milestones:", st.session_state["num_milestones"]) 
    
    elif current_step == 4:  # Milestone Input (Repeated for each milestone)
        chatbot_message(
            "Let's define the milestones for your project. Please provide the following details for each milestone:"
        )

        # if "milestone_index" not in st.session_state:
        #     st.session_state["milestone_index"] = 0

        # if "milestones" not in st.session_state:
        #     st.session_state["milestones"] = []

        # # Ensure num_milestones is initialized
        # if "num_milestones" in st.session_state and st.session_state["num_milestones"] > 0:
        #     st.session_state["num_milestones"] = 0

        i = st.session_state["milestone_index"]

        # Check if num_milestones is set and greater than 0
        if (
            "num_milestones" in st.session_state
            and st.session_state["num_milestones"] > 0
        ):
            if i < st.session_state["num_milestones"]:
                st.write(f"**Milestone {i+1}**")

                # Update existing milestone or create a new one
                if i < len(st.session_state["milestones"]):
                    # Update existing milestone
                    st.session_state["milestones"][i] = {
                        "Name": st.text_input(f"Milestone {i + 1} Name", key=f"milestone_{i+1}_name"),
                        "Start Date": st.date_input(f"Milestone {i + 1} Start Date", key=f"milestone_{i+1}_start_date"),
                        "End Date": st.date_input(f"Milestone {i + 1} End Date", key=f"milestone_{i+1}_end_date"),
                        "Milestone Owner": st.text_input(f"Milestone {i + 1} Owner", key=f"milestone_{i+1}_owner"),
                        "Progress": st.number_input(f"Milestone {i + 1} Progress (0-100)", min_value=0, max_value=100, step=1, value=0, key=f"milestone_{i+1}_progress"),
                        "Status": "" 
                    }
                else:
                    # Create a new milestone
                    st.session_state["milestones"].append({
                        "Name": st.text_input(f"Milestone {i + 1} Name", key=f"milestone_{i+1}_name"),
                        "Start Date": st.date_input(f"Milestone {i + 1} Start Date", key=f"milestone_{i+1}_start_date"),
                        "End Date": st.date_input(f"Milestone {i + 1} End Date", key=f"milestone_{i+1}_end_date"),
                        "Milestone Owner": st.text_input(f"Milestone {i + 1} Owner", key=f"milestone_{i+1}_owner"),
                        "Progress": st.number_input(f"Milestone {i + 1} Progress (0-100)", min_value=0, max_value=100, step=1, value=0, key=f"milestone_{i+1}_progress"),
                        "Status": "" 
                    })

                # "Add Another Milestone" button is only shown if more milestones are needed
                if st.session_state["milestone_index"] < st.session_state["num_milestones"] - 1:
                    if st.button("Add Another Milestone", on_click=go_to_step, args=(4,), key="add_milestone_button"):
                        st.session_state["milestone_index"] += 1

                # "Proceed" button is shown only when all milestones have been filled
                elif st.session_state["milestone_index"] == st.session_state["num_milestones"] - 1:
                    if st.button("Proceed", on_click=go_to_step, args=(5,), key="Proceed"):  # Proceed to Step 4.1
                        st.session_state["current_step"] = 5
                else:
                    st.warning("Please enter a positive number of milestones.")
        else:
            st.warning("Please enter the number of milestones first.")


    # elif current_step == 4.1:  # Prompt for Modifications
    #     chatbot_message("Do you want to make further modifications to your project plan?")
    #     col1, col2 = st.columns(2)
    #     with col1:
    #         if st.button("Yes", on_click=go_to_step, args=(5,), key="yes_modify_button"):
    #             st.session_state["current_step"] = 5
    #     with col2:
    #         if st.button("No", on_click=go_to_step, args=(6,), key="no_modify_button"):
    #             st.session_state["current_step"] = 6
    #             st.session_state["milestone_index"] = 0  # Reset milestone_index

    elif current_step == 5:  # Milestone Actions
        chatbot_message("Let's modify your project milestones.")  
        st.session_state["milestones"] = milestone_actions(st.session_state["milestones"],st.session_state["file_path"])

        # --- Add Dependency Question ---
        if st.button("Add Dependency", on_click=go_to_step, args=(8.1,), key="add_dependency_button"):  # Unique key
            st.session_state["current_step"] = 8.1

    # elif current_step == 5.1: # Dependency Question
    #     chatbot_message("Does this project plan have any dependencies on other projects?")
    #     col1, col2 = st.columns(2)
    #     with col1:
    #         if st.button("Yes", on_click=go_to_step, args=(8.1,), key="yes_dependency_button"):
    #             st.session_state["current_step"] = 8.1  # Dependency input
    #     with col2:
    #         if st.button("No", on_click=go_to_step, args=(6,), key="no_dependency_button"):
    #             st.session_state["current_step"] = 6  # Proceed without dependencies

    elif current_step == 8.1:  # Dependency Input
        handle_dependency_input(
            st.session_state["milestones"],
            st.session_state["project_name"],
            st.session_state["file_path"],
        )


    elif current_step == 6:  # Review Milestones
        chatbot_message("Would you like to review your milestones?")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Yes", on_click=go_to_step, args=(6.1,), key="yes_review_button"):
                st.session_state["current_step"] = 6.1
        with col2:
            if st.button("No", on_click=go_to_step, args=(12,), key="no_review_button"):
                st.session_state["current_step"] = 12

    elif current_step == 6.1:  # Review Milestones Display
        chatbot_message("Here are your current milestones:")
        display_project_details(
            st.session_state["project_name"],
            st.session_state["description"],
            st.session_state["stakeholders"],
            st.session_state["target_end_date"],
            st.session_state["milestones"],
        )
        chatbot_message("Would you like to make any changes to your milestones?")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Yes", on_click=go_to_step, args=(5,), key="yes_change_button"):
                st.session_state["current_step"] = 5
        with col2:
            if st.button("No", on_click=go_to_step, args=(7,), key="no_change_button"):
                st.session_state["current_step"] = 7


    elif current_step == 7: 
        if not st.session_state["file_path"]:
            chatbot_message("Great! Let's save your project plan. I'll create a file for you to download.")

            # Create the Excel file in memory
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add your Excel data to the worksheet
            ws["A1"] = "Project Name:"
            ws["B1"].font = Font(size=14, bold=True)
            ws["A2"] = "Description:"
            ws["A3"] = "Stakeholders:"
            ws["A4"] = "Target End Date:"
            ws["A6"] = "Milestone Name"
            ws["B6"] = "Start Date"
            ws["C6"] = "End Date"
            ws["D6"] = "Milestone Owner"
            ws["E6"] = "Progress"
            ws["F6"] = "Status"

            ws["B1"] = st.session_state["project_name"]
            ws["B2"] = st.session_state["description"]
            ws["B3"] = st.session_state["stakeholders"]
            ws["B4"] = st.session_state["target_end_date"].strftime("%Y-%m-%d") if st.session_state["target_end_date"] else "N/A"

            for i, milestone in enumerate(st.session_state["milestones"], start=7):
                ws[f"A{i}"] = milestone["Name"]
                ws[f"B{i}"] = milestone["Start Date"].strftime("%Y-%m-%d") if milestone["Start Date"] else "N/A"
                ws[f"C{i}"] = milestone["End Date"].strftime("%Y-%m-%d") if milestone["End Date"] else "N/A"
                ws[f"D{i}"] = milestone["Milestone Owner"]
                ws[f"E{i}"] = milestone["Progress"]
                ws[f"F{i}"] = milestone.get("Status", "")

            # Create a bytes stream to hold the Excel file
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Convert the bytes to a base64 encoded string
            b64_string = base64.b64encode(output.read()).decode("utf-8")

            # Create a download link
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_string}" download="{st.session_state["project_name"]}.xlsx">Download Project Plan</a>'

            st.markdown(href, unsafe_allow_html=True)

            # Add JavaScript to persist the current_step after download
            st.markdown(
                """
                <script>
                    window.addEventListener('load', function() {
                        // Get the current_step value from Streamlit session state (replace with your actual key)
                        var currentStep = JSON.parse(sessionStorage.getItem('streamlit_session_state'))['current_step']; 

                        // Store the current_step in the browser's local storage
                        localStorage.setItem('current_step', currentStep); 
                    });
                </script>
                """,
                unsafe_allow_html=True,
            )

            # Ask about analysis after download
            chatbot_message(
                "Your project plan is ready for download! Anything you would like to do as followings:"
            )
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Analyze it", on_click=go_to_step, args=(7.1,), key="analyze_button"):
                    st.session_state["current_step"] = 7.1
            with col2:
                if st.button("Provide Feedback", on_click=go_to_step, args=(13,), key="feedback_button"):
                    st.session_state["current_step"] = 13

    # # --- Milestone Actions ---
    # elif current_step == 5: 
    #     chatbot_message("Let's modify your projecct milestones.") # Milestone Actions
    #     st.session_state["milestones"] = milestone_actions(st.session_state["milestones"])
    #     st.button("Next", on_click=go_to_step, args=(6,))

            
    elif current_step == 7.1:  # Project Analysis (New Project or Loaded File)
        chatbot_message("Let's analyze your project plan!")

        # Calculate Program Duration
        (extreme_start_date, extreme_end_date, program_duration) = calculate_program_duration(
            st.session_state["milestones"]
        )

        chatbot_message(
            f"Your project is planned to run from {extreme_start_date.strftime('%Y-%m-%d')} to {extreme_end_date.strftime('%Y-%m-%d')}, a total of {program_duration} days."
        )

        # --- COLLECT DELAY WARNINGS ---
        delay_warnings = []  # Store warnings in a list

        # Analyze Milestone Progress
        milestone_data = []
        for milestone in st.session_state["milestones"]:
            status, warning, expected_progress = analyze_milestone_progress(
                milestone,
                extreme_start_date,
                extreme_end_date,
                st.session_state["target_end_date"],
            )
            
            # --- ONLY ADD WARNINGS FOR SIGNIFICANT DELAYS --- 
            if status == "Behind Schedule":
                milestone_end = milestone["End Date"] 
                milestone_end = milestone_end.date() if isinstance(milestone_end, datetime) else milestone_end

                today = datetime.now().date()
                days_behind = (today - milestone_end).days

                if days_behind >= 7: # Example: Consider a milestone delayed if it's 7 or more days past due
                    delay_warnings.append(warning)

            milestone_data.append(
                {
                    "Name": milestone["Name"],
                    "Status": status,
                    # --- Format Expected and Actual Progress --- 
                    "Expected Progress": f"{expected_progress:.2f}", 
                    "Actual Progress": f"{milestone['Progress']:.2f}",
                    "Start Date": milestone["Start Date"],
                    "End Date": milestone["End Date"]
                }
            )

        # --- DISPLAY DELAY WARNINGS IN A BOX ---
        if delay_warnings:
            st.warning("Project Warnings:")
            with st.expander("View Milestone Delay Details", expanded=False):  # Expandable box
                for warning in delay_warnings:
                    st.markdown(f"- {warning}")  # Display each warning with a bullet point

        # Display Milestone Status
        st.subheader("Milestone Status:")
        st.table(milestone_data)

        # --- Program Plan Visualization ---
        st.subheader("Program Plan Visualization:")

         # --- CREATE THE GANTT CHART DATA --- (This is where you add the code)
        df_gantt = pd.DataFrame(
            {
                "Task": [m["Name"] for m in st.session_state["milestones"]],
                "Start": [m["Start Date"] for m in st.session_state["milestones"]],
                "Finish": [m["End Date"] for m in st.session_state["milestones"]],
                # --- Convert Actual Progress to float --- (Corrected code)
                "Complete": [float(m["Actual Progress"]) / 100 for m in milestone_data],  # For color 
            }
        )

        # --- DETERMINE MILESTONE COLORS ---
        colors = []
        today = datetime.now().date()
        for m in milestone_data:
            start_date = m["Start Date"].date() if isinstance(m["Start Date"], datetime) else m["Start Date"]
            end_date = m["End Date"].date() if isinstance(m["End Date"], datetime) else m["End Date"]

            # 1. Delayed Milestones (Red)
            if m["Status"] == "Behind Schedule": 
                colors.append("red") 

            # 2. Completed Milestones (Green)
            elif end_date < today and m["Actual Progress"] == 100: 
                colors.append("green")  

            # 3. In-Progress Milestones (Light Blue)
            elif start_date <= today <= end_date: 
                colors.append("lightblue")  

            # 4. Future Milestones (Gray)
            else:
                colors.append("gray")  

        # --- CREATE THE GANTT CHART FIGURE ---
        fig = px.timeline(
            df_gantt, 
            x_start="Start", 
            x_end="Finish", 
            y="Task",
            color="Complete",
            color_continuous_scale=[(0, "red"), (0.99, "lightblue"), (1, "green")], 
            range_color=(0, 1),
        )

        # --- UPDATE PLOTLY LAYOUT ---
        fig.update_yaxes(autorange="reversed") 
        fig.update_layout(
            title="Program Plan",
            xaxis_title="Date",
            yaxis_title="Milestones",
            showlegend=False,
            xaxis=dict(tickformat="%Y-%m-%d"),
            hovermode="x unified",

            # --- VISUAL TWEAKS ---
            plot_bgcolor='rgba(0,0,0,0)' # Transparent background
        )

        # --- ADD TODAY'S DATE INDICATOR ---
        fig.add_vline(x=today, line_width=3, line_dash="dash", line_color="red") # Make it more visible

        # --- DISPLAY THE CHART ---
        st.plotly_chart(fig, use_container_width=True)

        # Milestone Progress Visualization (Plotly)
        display_plotly_progress(milestone_data)

        st.button("Next", on_click=go_to_step, args=(8,), key="next_button_7_1")

    elif current_step == 8:  # After Analysis
        chatbot_message("Analysis complete! What would you like to do next?")
        st.button("Provide Feedback", on_click=go_to_step, args=(13,), key="feedback_button_8")


    elif current_step == 8.1:  # Dependency Input
        handle_dependency_input(
            st.session_state["milestones"],
            st.session_state["project_name"],
            st.session_state["file_path"],
        )

    elif current_step == 9:  # Dependency Addition
        chatbot_message(
            "I've added a dependency milestone to your project. Would you like to give feedback about the chatbot?"
        )
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Yes", on_click=go_to_step, args=(10,)):
                st.session_state["current_step"] = 10  # Proceed to Step 10
        with col2:
            if st.button("No", on_click=go_to_step, args=(0,)):
                st.session_state["current_step"] = 0  # Return to Step 1

    elif current_step == 10:  # Feedback
        chatbot_message(
            "Thanks for the feedback! On a scale of 1 to 5, how would you rate your experience with this chatbot?"
        )
        st.session_state["rating"] = st.slider(
            "Rating (1-5)", min_value=1, max_value=5, value=3
        )
        st.session_state["suggestions"] = st.text_area("Your Suggestions")
        if st.button("Submit Feedback", on_click=go_to_step, args=(0,)):
            chatbot_message(
                "Thank you for your feedback! It helps us improve."
            )
            # chatbot_message(
            #     "Would you like to: \n Create a new project plan \n Analyze another existing project plan \n Modify another existing project plan \n Exit"
            # )
            st.session_state["current_step"] = 0  # Return to Step 1

    elif current_step == 11:  # Modify Existing Project 
        chatbot_message("Great! Let's modify your plan. Please upload your project plan (.xlsx file).")
        uploaded_file = st.file_uploader("Choose a file", type=["xlsx"])

        if uploaded_file is not None:
            try:
                # --- LOAD PROJECT PLAN ---
                (
                    st.session_state["project_name"],
                    st.session_state["description"],
                    st.session_state["stakeholders"],
                    st.session_state["target_end_date"],
                    st.session_state["milestones"],
                ) = load_project_plan(uploaded_file.name)
                st.session_state["file_path"] = uploaded_file.name

                # --- GO TO MILESTONE MODIFICATION SECTION ---
                if st.button("Let's Modify it!", on_click=go_to_step, args=(5,), key="next_button_11"):

                    st.session_state["current_step"] = 5  # Go to Milestone Modification

            except Exception as e:
                st.error(f"Error loading project: {e}")

    

    elif current_step == 12:  # After Modify, before Dependency
        # Automatically save the modified project plan
        if st.session_state["file_path"]:
            try:
                save_project_plan(
                    st.session_state["project_name"],
                    st.session_state["description"],
                    st.session_state["stakeholders"],
                    st.session_state["target_end_date"],
                    st.session_state["milestones"],
                    st.session_state["file_path"],  # Use the original file name
                )
                st.success(
                    f"Project plan saved to '{st.session_state['file_path']}'!"
                )

                # --- Update Chatbot Messages ---
                chatbot_message("Your modified project plan has been saved!") 
                chatbot_message("What would you like to do next?")
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("Analyze Modified Plan", on_click=go_to_step, args=(11.1,), key="analyze_button_12"):
                        st.session_state["current_step"] = 11.1 # Analyze modified plan
                with col2:
                    if st.button("Provide Feedback", on_click=go_to_step, args=(13,), key="feedback_button_12"):
                        st.session_state["current_step"] = 13     # Feedback 
            except Exception as e:
                st.error(f"Error saving: {e}")
        else:
            st.error("No file path found. Please upload a file first.")
            st.session_state["current_step"] = 11  # Go back to file upload

    

    elif current_step == 11.1: # Analysis after Modification
        chatbot_message("Let's analyze your modified project plan!")

        # --- CALCULATE MILESTONE DATA ---
        milestone_data = []
        (extreme_start_date, extreme_end_date, program_duration) = calculate_program_duration(st.session_state["milestones"])
        
        delay_warnings = []
        for milestone in st.session_state["milestones"]:
            status, warning, expected_progress = analyze_milestone_progress(
                milestone,
                extreme_start_date,
                extreme_end_date,
                st.session_state["target_end_date"],
            )
            
            # --- ONLY ADD WARNINGS FOR SIGNIFICANT DELAYS --- 
            if status == "Behind Schedule":
                milestone_end = milestone["End Date"] 
                milestone_end = milestone_end.date() if isinstance(milestone_end, datetime) else milestone_end
    
                today = datetime.now().date()
                days_behind = (today - milestone_end).days
                
                if days_behind >= 7: # Example: Consider a milestone delayed if it's 7 or more days past due
                    delay_warnings.append(warning)


            milestone_data.append(
                {
                    "Name": milestone["Name"],
                    "Status": status,
                    "Expected Progress": expected_progress,
                    "Actual Progress": milestone["Progress"],
                    "Start Date": milestone["Start Date"],
                    "End Date": milestone["End Date"]
                }
            )

        # --- DISPLAY PROJECT DURATION ---
        chatbot_message(
            f"Your project is planned to run from {extreme_start_date.strftime('%Y-%m-%d')} to {extreme_end_date.strftime('%Y-%m-%d')}, a total of {program_duration} days."
        )

        # --- DISPLAY DELAY WARNINGS ---
        if delay_warnings:
            st.warning("Project Warnings:")
            with st.expander("View Milestone Delay Details", expanded=False):  
                for warning in delay_warnings:
                    st.markdown(f"- {warning}") 

        # --- DISPLAY MILESTONE STATUS TABLE ---
        st.subheader("Milestone Status:")
        st.table(milestone_data)

        # --- Program Plan Visualization ---
        st.subheader("Program Plan Visualization:")

        # --- Create Plotly Gantt Chart Data ---
        df_gantt = pd.DataFrame(
            {
                "Task": [m["Name"] for m in st.session_state["milestones"]],
                "Start": [m["Start Date"] for m in st.session_state["milestones"]],
                "Finish": [m["End Date"] for m in st.session_state["milestones"]],
                # --- Convert Actual Progress to float --- (Corrected code)
                "Complete": [float(m["Actual Progress"]) / 100 for m in milestone_data],  # For color 
            }
        )


        # --- DETERMINE MILESTONE COLORS ---
        colors = []
        today = datetime.now().date()
        for m in milestone_data:
            start_date = m["Start Date"].date() if isinstance(m["Start Date"], datetime) else m["Start Date"]
            end_date = m["End Date"].date() if isinstance(m["End Date"], datetime) else m["End Date"]

            # 1. Delayed Milestones (Red)
            if m["Status"] == "Behind Schedule": 
                colors.append("red") 

            # 2. Completed Milestones (Green)
            elif end_date < today and m["Actual Progress"] == 100: 
                colors.append("green")  

            # 3. In-Progress Milestones (Light Blue)
            elif start_date <= today <= end_date: 
                colors.append("lightblue")  

            # 4. Future Milestones (Gray)
            else:
                colors.append("gray")  

        # --- CREATE THE GANTT CHART FIGURE ---
        fig = px.timeline(
            df_gantt, 
            x_start="Start", 
            x_end="Finish", 
            y="Task",
            color="Complete",
            color_continuous_scale=[(0, "red"), (0.99, "lightblue"), (1, "green")], 
            range_color=(0, 1),
        )

        # --- UPDATE PLOTLY LAYOUT ---
        fig.update_yaxes(autorange="reversed") 
        fig.update_layout(
            title="Program Plan",
            xaxis_title="Date",
            yaxis_title="Milestones",
            showlegend=False,
            xaxis=dict(tickformat="%Y-%m-%d"),
            hovermode="x unified",

            # --- VISUAL TWEAKS ---
            plot_bgcolor='rgba(0,0,0,0)' # Transparent background
        )

        # --- ADD TODAY'S DATE INDICATOR ---
        fig.add_vline(x=today, line_width=3, line_dash="dash", line_color="red") # Make it more visible

        # --- DISPLAY THE CHART ---
        st.plotly_chart(fig, use_container_width=True)

        # --- MILESTONE PROGRESS BAR VISUALIZATION ---
        display_plotly_progress(milestone_data)

        # --- GO TO FEEDBACK BUTTON ---
        st.button("Go to Feedback", on_click=go_to_step, args=(13,), key="feedback_button_11_1") 


    elif current_step == 13:  # User Feedback (Modified Project)
        chatbot_message(
            "Thanks! How would you rate your experience with this chatbot on a scale of 1 to 5?"
        )
        st.session_state["rating"] = st.slider(
            "Rating (1-5)", min_value=1, max_value=5, value=3
        )
        st.session_state["suggestions"] = st.text_area("Your Suggestions")
        if st.button("Submit Feedback", on_click=go_to_step, args=(0,)):
            chatbot_message(
                "Thank you for your feedback! It helps us improve."
            )
            # chatbot_message(
            #     "Would you like to: \n Create a new project plan \n Analyze another existing project plan \n Modify another existing project plan \n Exit"
            # )
            st.session_state["current_step"] = 0  # Return to Step 1

    

chatbot_flow()  # Run the chatbot flow